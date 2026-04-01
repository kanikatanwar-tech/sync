import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import re
from collections import defaultdict

st.set_page_config(
    page_title="Timetable Sync & Conflict Tool",
    page_icon="📅",
    layout="wide"
)

# ─────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────
DAYS    = ["MON", "TUE", "WED", "THU", "FRI", "SAT"]
PERIODS = ["P0", "P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8"]
TOTAL_SLOTS = 54

WORKLOAD_THRESHOLDS = {"Overloaded": 80, "High Load": 60, "Normal": 35, "Low Load": 0}

NON_TEACHING_KEYWORDS = {
    'NEEEV', 'CT', 'KB', 'MDM', 'EVGC', 'LIB', 'LIBRARY', 'GAMES',
}

HDR_FILL  = PatternFill("solid", fgColor="1F3864")
HDR_FONT  = Font(color="FFFFFF", bold=True, name="Arial", size=10)
TTL_FILL  = PatternFill("solid", fgColor="2E75B6")
TTL_FONT  = Font(color="FFFFFF", bold=True, name="Arial", size=11)
DAY_FILL  = PatternFill("solid", fgColor="D6E4F0")
DAY_FONT  = Font(bold=True, name="Arial", size=10)
SUM_FILL  = PatternFill("solid", fgColor="E2EFDA")
NRM_FONT  = Font(name="Arial", size=9)
NRM_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CTR_ALIGN = Alignment(horizontal="center", vertical="center")

CONFLICT_FILLS = {
    "CRITICAL": PatternFill("solid", fgColor="FF0000"),
    "WARNING":  PatternFill("solid", fgColor="FF9900"),
    "INFO":     PatternFill("solid", fgColor="BDD7EE"),
}
CONFLICT_FONTS = {
    "CRITICAL": Font(color="FFFFFF", bold=True, name="Arial", size=9),
    "WARNING":  Font(color="000000", bold=True, name="Arial", size=9),
    "INFO":     Font(color="000000", bold=True, name="Arial", size=9),
}

STATUS_COLORS = {
    "Overloaded": "FF0000", "High Load": "FF9900",
    "Normal": "00B050",     "Low Load": "0070C0",
}

SEV_EMOJI = {"CRITICAL": "🔴", "WARNING": "🟠", "INFO": "🔵"}
SEV_COLOR = {"CRITICAL": "#FF4444", "WARNING": "#FF9900", "INFO": "#0088CC"}

CONFLICT_DESCRIPTIONS = {
    "C1":  "Teacher assigned to multiple classes in the same slot",
    "C2":  "Two teachers assigned for the same class+subject simultaneously",
    "C3":  "Parallel/split sessions for same class in same slot (verify if intentional)",
    "C4":  "Same subject scheduled 3+ times in one day for a class",
    "C5":  "Workload anomaly — too high or mathematically impossible",
    "C6":  "Teacher has zero slots assigned",
    "C7":  "Class has no teaching on an entire day",
    "C8":  "Same teacher name appears in multiple sheets (duplicate entry)",
    "C10": "Class has 5+ consecutive teaching periods without any break",
}


def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_cell(cell, value=None, fill=None, font=None, align=None):
    if value is not None: cell.value = value
    if fill:  cell.fill  = fill
    if font:  cell.font  = font
    if align: cell.alignment = align


# ─────────────────────────────────────────────
#  CELL VALUE PARSER
# ─────────────────────────────────────────────
def parse_cell_value(cell_val):
    val = cell_val.strip()
    class_pattern = r'(\d{1,2}[A-F](?:\+\d{1,2}[A-F])*)'
    m = re.match(r'^' + class_pattern + r'\s*(?:-\s*)?(.+)$', val)
    if m:
        class_part = m.group(1)
        subject    = m.group(2).strip().lstrip('-').strip()
        classes    = [c.strip() for c in class_part.split('+')]
        return classes, subject
    return [], val


def is_non_teaching(subject):
    s = subject.upper().strip()
    if s in NON_TEACHING_KEYWORDS:
        return True
    if s.startswith('RN OR') or s.startswith('SOL OR') or s.startswith('RN '):
        return True
    return False


def is_parallel_slot(subject):
    s = subject.upper().strip()
    return s.startswith('SOL OR') or s.startswith('RN OR') or s.startswith('RN ')


def class_sort_key(c):
    m = re.match(r'(\d+)([A-Z])', str(c))
    return (int(m.group(1)), m.group(2)) if m else (999, str(c))


# ─────────────────────────────────────────────
#  PARSE UNAVAILABILITY SHEET
# ─────────────────────────────────────────────
def parse_unavailability(file_bytes):
    """
    Returns:
      unavailability: dict  {sheet_name -> list of unavailable period strings e.g. ['0','1','2']}
      unavail_raw_wb: openpyxl Worksheet  (for copying as-is into Summary Report)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    unavailability = {}   # keyed by SHEET NAME (matched to teacher_info keys)
    unavail_rows   = []   # raw rows for re-use in writing

    if "Unavailability" not in wb.sheetnames:
        return unavailability, unavail_rows, wb

    ws = wb["Unavailability"]
    rows = list(ws.iter_rows(values_only=True))

    for row in rows[1:]:   # skip header row
        if not row[0]:
            continue
        teacher_key = str(row[0]).strip()   # e.g. "MEENAKSHI GUPTA-20132491"
        periods_str = str(row[1]).strip() if row[1] else ""
        periods     = [p.strip() for p in periods_str.split(',') if p.strip()]
        unavailability[teacher_key] = periods
        unavail_rows.append({"teacher_key": teacher_key, "periods": periods})

    return unavailability, unavail_rows, wb


# ─────────────────────────────────────────────
#  PARSE TEACHER TT
# ─────────────────────────────────────────────
def parse_teacher_tt(file_bytes):
    wb   = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    teacher_data     = {}
    teacher_info     = {}
    teaching_summary = {}

    for sheet_name in wb.sheetnames:
        if sheet_name == "Unavailability":
            continue                      # skip — parsed separately

        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        raw_name = rows[1][0] if len(rows) > 1 else sheet_name
        name_str = str(raw_name) if raw_name else sheet_name
        match    = re.match(r'^(.+?)-(\S+)$', name_str.strip())
        tname, tid = (match.group(1).strip(), match.group(2).strip()) if match else (name_str.strip(), "")
        teacher_info[sheet_name] = {"name": tname, "id": tid}

        grid = {}
        for day_idx, day in enumerate(DAYS):
            row_idx = 3 + day_idx
            if row_idx >= len(rows):
                break
            row = rows[row_idx]
            for p_idx, period in enumerate(PERIODS):
                col_idx = 1 + p_idx
                val = row[col_idx] if col_idx < len(row) else None
                if val and str(val).strip() and str(val).strip().lower() != 'nan':
                    grid[(day, period)] = str(val).strip()
                else:
                    grid[(day, period)] = ""
        teacher_data[sheet_name] = grid

        summary_rows = []
        in_summary   = False
        for row in rows:
            if row[0] and str(row[0]).strip().upper() == "TEACHING SUMMARY":
                in_summary = True
                continue
            if in_summary:
                cls, subj, periods = row[0], row[1], row[2]
                if cls is None and subj and str(subj).strip().upper() == "TOTAL":
                    break
                if cls and subj and periods is not None:
                    try:
                        p = int(float(str(periods)))
                    except Exception:
                        p = 0
                    summary_rows.append({"Class": str(cls).strip(), "Subject": str(subj).strip(), "Periods": p})
        teaching_summary[sheet_name] = summary_rows

    return teacher_data, teacher_info, teaching_summary


# ─────────────────────────────────────────────
#  BUILD CLASS MAPPING
# ─────────────────────────────────────────────
def build_class_mapping(teacher_data, teacher_info):
    class_tt = {}
    for sheet_name, grid in teacher_data.items():
        tname = teacher_info[sheet_name]["name"]
        for (day, period), cell_val in grid.items():
            if not cell_val:
                continue
            classes, subject = parse_cell_value(cell_val)
            for cls in classes:
                if cls not in class_tt:
                    class_tt[cls] = {}
                key     = (day, period)
                display = f"{subject} ({tname})"
                if key in class_tt[cls]:
                    if tname not in class_tt[cls][key]:
                        class_tt[cls][key] = class_tt[cls][key] + f"\n{display}"
                else:
                    class_tt[cls][key] = display
    return class_tt


# ─────────────────────────────────────────────
#  WORKLOAD & BLUEPRINT
# ─────────────────────────────────────────────
def compute_workload(teacher_data, teacher_info, teaching_summary, unavailability):
    """
    unavailability: dict {teacher_sheet_key -> list of unavail period strings ['0','1','2']}
    Available slots = 54 - (len(unavail_periods) × 6 days)
    Workload % = periods_allotted / available_slots × 100
    """
    rows = []
    for sheet_name, grid in teacher_data.items():
        info    = teacher_info[sheet_name]
        summary = teaching_summary.get(sheet_name, [])
        total   = sum(r["Periods"] for r in summary)
        if total == 0:
            total = sum(1 for v in grid.values() if v)

        # Match unavailability by sheet name (exact key) or by name-id pattern
        unavail_periods = unavailability.get(sheet_name, [])
        # Also try matching by teacher full name+id if direct key not found
        if not unavail_periods:
            tname_id = f"{info['name']}-{info['id']}"
            for ukey, uperiods in unavailability.items():
                if ukey.upper() == tname_id.upper() or ukey.upper() == info['name'].upper():
                    unavail_periods = uperiods
                    break

        n_unavail_slots = len(unavail_periods) * len(DAYS)   # same periods every day
        available_slots = TOTAL_SLOTS - n_unavail_slots
        available_slots = max(available_slots, 1)             # guard against div/0

        pct    = (total / available_slots) * 100
        status = next((s for s, t in sorted(WORKLOAD_THRESHOLDS.items(), key=lambda x: -x[1]) if pct >= t), "Low Load")

        rows.append({
            "Teacher Name":      info["name"],
            "Teacher ID":        info["id"],
            "Periods Allotted":  total,
            "Unavail Periods":   ", ".join(unavail_periods) if unavail_periods else "None",
            "Available Slots":   available_slots,
            "Workload %":        round(pct, 2),
            "Status":            status,
        })
    return rows


def compute_blueprint(teacher_data, teacher_info):
    blueprint = {}
    for sheet_name, grid in teacher_data.items():
        counts = {}
        for (day, period), cell_val in grid.items():
            if not cell_val:
                continue
            classes, subject = parse_cell_value(cell_val)
            for cls in classes:
                counts[(cls, subject)] = counts.get((cls, subject), 0) + 1
        blueprint[sheet_name] = {"info": teacher_info[sheet_name], "counts": counts}
    return blueprint


# ─────────────────────────────────────────────
#  CONFLICT DETECTION ENGINE
# ─────────────────────────────────────────────
def detect_conflicts(teacher_data, teacher_info):
    conflicts = []

    teacher_slot_assignments = defaultdict(list)
    class_slot_assignments   = defaultdict(list)

    for sheet, grid in teacher_data.items():
        info = teacher_info[sheet]
        for (day, period), cell_val in grid.items():
            if not cell_val:
                continue
            classes, subject = parse_cell_value(cell_val)
            teacher_slot_assignments[(sheet, day, period)].append(
                {"classes": classes, "subject": subject, "raw": cell_val}
            )
            for cls in classes:
                class_slot_assignments[(cls, day, period)].append(
                    {"teacher": info["name"], "tid": info["id"],
                     "subject": subject, "raw": cell_val, "sheet": sheet}
                )

    # C1: Teacher assigned to multiple classes in same slot
    for (sheet, day, period), entries in teacher_slot_assignments.items():
        info    = teacher_info[sheet]
        teaching = [e for e in entries if not is_non_teaching(e["subject"])]
        if len(teaching) > 1:
            raws = [e["raw"] for e in teaching]
            conflicts.append({
                "severity": "CRITICAL", "conflict_id": "C1",
                "title": f"Teacher double-booked: {info['name']} on {day} {period}",
                "detail": (f"Teacher **{info['name']}** is assigned to **multiple classes** "
                           f"on **{day} {period}**: {', '.join(raws)}. "
                           f"A teacher cannot be in two places at once."),
                "day": day, "period": period, "teacher": info["name"], "cls": None,
            })

    # C2 & C3: Class slot analysis
    for (cls, day, period), entries in class_slot_assignments.items():
        teaching = [e for e in entries
                    if not is_non_teaching(e["subject"]) and not is_parallel_slot(e["subject"])]
        if len(teaching) <= 1:
            continue
        by_subject = defaultdict(list)
        for e in teaching:
            by_subject[e["subject"]].append(e)

        # C2: Same subject, multiple teachers
        for subj, es in by_subject.items():
            if len(es) > 1:
                conflicts.append({
                    "severity": "CRITICAL", "conflict_id": "C2",
                    "title": f"Class {cls}: 2 teachers for '{subj}' on {day} {period}",
                    "detail": (f"Class **{cls}** is assigned **{subj}** by "
                               f"**{' and '.join(e['teacher'] for e in es)}** "
                               f"at the same time (**{day} {period}**). Only one teacher should cover this."),
                    "day": day, "period": period,
                    "teacher": " / ".join(e["teacher"] for e in es), "cls": cls,
                })

        # C3: Different subjects (elective split — INFO)
        unique_subjects = set(e["subject"] for e in teaching)
        if len(unique_subjects) > 1:
            conflicts.append({
                "severity": "INFO", "conflict_id": "C3",
                "title": f"Class {cls}: parallel sessions on {day} {period}",
                "detail": (f"Class **{cls}** has **parallel sessions** on **{day} {period}** "
                           f"with different subjects: **{' / '.join(sorted(unique_subjects))}**. "
                           f"Teachers: {', '.join(e['teacher'] for e in teaching)}. "
                           f"This is valid for elective groups — verify it's intentional."),
                "day": day, "period": period,
                "teacher": ", ".join(e["teacher"] for e in teaching), "cls": cls,
            })

    # C4: Same subject 3+ times on same day for same class
    subj_day_slots = defaultdict(list)
    for (cls, day, period), entries in class_slot_assignments.items():
        for e in entries:
            if not is_non_teaching(e["subject"]) and not is_parallel_slot(e["subject"]):
                subj_day_slots[(cls, day, e["subject"])].append(period)

    for (cls, day, subj), periods_list in subj_day_slots.items():
        if len(periods_list) > 2:
            conflicts.append({
                "severity": "WARNING", "conflict_id": "C4",
                "title": f"Class {cls}: '{subj}' appears {len(periods_list)}x on {day}",
                "detail": (f"Class **{cls}** has **{subj}** scheduled **{len(periods_list)} times** "
                           f"on **{day}** (periods: {', '.join(sorted(periods_list))}). "
                           f"More than 2 periods of the same subject in one day is usually excessive."),
                "day": day, "period": ", ".join(sorted(periods_list)),
                "teacher": None, "cls": cls,
            })

    # C5: Workload anomalies
    for sheet, grid in teacher_data.items():
        info  = teacher_info[sheet]
        total = sum(1 for v in grid.values() if v)
        if total > TOTAL_SLOTS:
            conflicts.append({
                "severity": "CRITICAL", "conflict_id": "C5",
                "title": f"Impossible workload: {info['name']} ({total} slots > {TOTAL_SLOTS} max)",
                "detail": (f"Teacher **{info['name']}** has **{total} filled slots** which exceeds "
                           f"the maximum of **{TOTAL_SLOTS}** (6 days × 9 periods). "
                           f"This indicates a data entry error — check for duplicate rows."),
                "day": None, "period": None, "teacher": info["name"], "cls": None,
            })
        elif total > 48:
            conflicts.append({
                "severity": "WARNING", "conflict_id": "C5",
                "title": f"Extreme workload: {info['name']} ({total}/{TOTAL_SLOTS} slots, {round(total/TOTAL_SLOTS*100,1)}%)",
                "detail": (f"Teacher **{info['name']}** has **{total}/{TOTAL_SLOTS}** slots filled. "
                           f"This is exceptionally high — please verify it's intentional."),
                "day": None, "period": None, "teacher": info["name"], "cls": None,
            })

    # C6: Teacher with zero slots
    for sheet, grid in teacher_data.items():
        info  = teacher_info[sheet]
        total = sum(1 for v in grid.values() if v)
        if total == 0:
            conflicts.append({
                "severity": "WARNING", "conflict_id": "C6",
                "title": f"No slots assigned: {info['name']}",
                "detail": (f"Teacher **{info['name']}** (ID: {info['id']}) has **no slots** "
                           f"assigned in the timetable. Verify this teacher should be in the file."),
                "day": None, "period": None, "teacher": info["name"], "cls": None,
            })

    # C7: Class has no coverage on an entire day
    class_day_coverage = defaultdict(set)
    for (cls, day, period), entries in class_slot_assignments.items():
        real = [e for e in entries if not is_non_teaching(e["subject"])]
        if real:
            class_day_coverage[cls].add(day)

    for cls in set(k[0] for k in class_slot_assignments):
        missing = [d for d in DAYS if d not in class_day_coverage[cls]]
        if missing:
            conflicts.append({
                "severity": "WARNING", "conflict_id": "C7",
                "title": f"Class {cls}: no teaching periods on {', '.join(missing)}",
                "detail": (f"Class **{cls}** has **no teaching slots** on **{', '.join(missing)}**. "
                           f"Verify this is intentional (e.g. holiday/games day) or a missing entry."),
                "day": ", ".join(missing), "period": None, "teacher": None, "cls": cls,
            })

    # C8: Same teacher name in multiple sheets
    name_to_sheets = defaultdict(list)
    for sheet, info in teacher_info.items():
        name_to_sheets[info["name"].upper()].append(sheet)
    for tname_upper, sheets in name_to_sheets.items():
        if len(sheets) > 1:
            tname = teacher_info[sheets[0]]["name"]
            conflicts.append({
                "severity": "CRITICAL", "conflict_id": "C8",
                "title": f"Duplicate teacher: '{tname}' appears in {len(sheets)} sheets",
                "detail": (f"Teacher **{tname}** has entries in **multiple sheets**: "
                           f"{', '.join(sheets)}. Each teacher should have exactly one sheet. "
                           f"Merge or remove the duplicate."),
                "day": None, "period": None, "teacher": tname, "cls": None,
            })

    # C10: Class with 5+ consecutive teaching periods
    all_classes = set(k[0] for k in class_slot_assignments)
    for cls in all_classes:
        for day in DAYS:
            day_slots = []
            for period in PERIODS:
                entries = class_slot_assignments.get((cls, day, period), [])
                real    = [e for e in entries
                           if not is_non_teaching(e["subject"]) and not is_parallel_slot(e["subject"])]
                day_slots.append(bool(real))

            consecutive = max_consec = 0
            for filled in day_slots:
                if filled:
                    consecutive += 1
                    max_consec   = max(max_consec, consecutive)
                else:
                    consecutive  = 0

            if max_consec >= 5:
                conflicts.append({
                    "severity": "WARNING", "conflict_id": "C10",
                    "title": f"Class {cls}: {max_consec} consecutive periods on {day} (no break)",
                    "detail": (f"Class **{cls}** has **{max_consec} consecutive teaching periods** "
                               f"on **{day}** with no free period. Consider adding a break."),
                    "day": day, "period": None, "teacher": None, "cls": cls,
                })

    return conflicts


# ─────────────────────────────────────────────
#  WRITE CONFLICT REPORT EXCEL
# ─────────────────────────────────────────────
def write_conflict_report(conflicts):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conflict Report"

    ws.merge_cells("A1:G1")
    apply_cell(ws["A1"], "TIMETABLE CONFLICT & ERROR REPORT", HDR_FILL, HDR_FONT, CTR_ALIGN)

    n_critical = sum(1 for c in conflicts if c["severity"] == "CRITICAL")
    n_warning  = sum(1 for c in conflicts if c["severity"] == "WARNING")
    n_info     = sum(1 for c in conflicts if c["severity"] == "INFO")

    ws.merge_cells("A2:G2")
    ws["A2"].value = (f"Total Issues: {len(conflicts)}   |   "
                      f"CRITICAL: {n_critical}   |   WARNING: {n_warning}   |   INFO: {n_info}")
    ws["A2"].font      = Font(bold=True, name="Arial", size=10)
    ws["A2"].alignment = CTR_ALIGN

    headers = ["Severity", "Code", "Title", "Day", "Period", "Teacher", "Details"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
        cell.alignment = CTR_ALIGN

    sev_order = {"CRITICAL": 0, "WARNING": 1, "INFO": 2}
    for ri, c in enumerate(sorted(conflicts, key=lambda x: sev_order.get(x["severity"], 3)), 4):
        sev  = c["severity"]
        fill = CONFLICT_FILLS[sev]
        font = CONFLICT_FONTS[sev]

        sev_cell = ws.cell(row=ri, column=1, value=sev)
        sev_cell.fill = fill
        sev_cell.font = font
        sev_cell.alignment = CTR_ALIGN

        ws.cell(row=ri, column=2, value=c["conflict_id"]).alignment = CTR_ALIGN
        ws.cell(row=ri, column=2).font = NRM_FONT

        tc = ws.cell(row=ri, column=3, value=c["title"])
        color_map = {"CRITICAL": "CC0000", "WARNING": "CC6600", "INFO": "0070C0"}
        tc.font = Font(bold=(sev != "INFO"), name="Arial", size=9, color=color_map[sev])

        ws.cell(row=ri, column=4, value=c.get("day") or "—").font = NRM_FONT
        ws.cell(row=ri, column=4).alignment = CTR_ALIGN
        ws.cell(row=ri, column=5, value=c.get("period") or "—").font = NRM_FONT
        ws.cell(row=ri, column=5).alignment = CTR_ALIGN
        ws.cell(row=ri, column=6, value=c.get("teacher") or "—").font = NRM_FONT
        ws.cell(row=ri, column=6).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        detail = re.sub(r'\*\*(.+?)\*\*', r'\1', c.get("detail", ""))
        cls_val = c.get("cls") or ""
        det_str = f"[Class: {cls_val}] {detail}".strip() if cls_val else detail
        dc = ws.cell(row=ri, column=7, value=det_str)
        dc.font = NRM_FONT
        dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for ci in range(1, 8):
            ws.cell(row=ri, column=ci).border = thin_border()

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 65

    for ri in range(4, 4 + len(conflicts)):
        ws.row_dimensions[ri].height = 32

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ─────────────────────────────────────────────
#  WRITE CLASS TT EXCEL
# ─────────────────────────────────────────────
def write_class_tt(class_tt, teacher_info):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for cls in sorted(class_tt.keys(), key=class_sort_key):
        ws   = wb.create_sheet(title=cls)
        grid = class_tt[cls]
        ws.merge_cells("A1:J1")
        apply_cell(ws["A1"], "CLASS TIMETABLE", HDR_FILL, HDR_FONT, CTR_ALIGN)
        ws.merge_cells("A2:J2")
        apply_cell(ws["A2"], f"Class: {cls}", TTL_FILL, TTL_FONT, CTR_ALIGN)
        for ci, h in enumerate(["Day"] + PERIODS, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CTR_ALIGN
        for ri, day in enumerate(DAYS, 4):
            ws.cell(row=ri, column=1, value=day).fill = DAY_FILL
            ws.cell(row=ri, column=1).font = DAY_FONT
            ws.cell(row=ri, column=1).alignment = CTR_ALIGN
            for pi, period in enumerate(PERIODS, 2):
                val  = grid.get((day, period), "")
                cell = ws.cell(row=ri, column=pi, value=val)
                cell.font = NRM_FONT
                cell.alignment = NRM_ALIGN
                cell.border = thin_border()
        ws.column_dimensions["A"].width = 8
        for ci in range(2, 11):
            ws.column_dimensions[get_column_letter(ci)].width = 28
        for r in range(4, 10):
            ws.row_dimensions[r].height = 40
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ─────────────────────────────────────────────
#  WRITE SUMMARY REPORT EXCEL
# ─────────────────────────────────────────────
def write_summary_report(teacher_data, teacher_info, class_tt, workload_rows, blueprint, unavail_rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_unavailability_sheet(wb, unavail_rows)   # first sheet
    _write_teacher_tt_summary(wb, teacher_data, teacher_info)
    _write_class_tt_summary(wb, class_tt)
    _write_workload_summary(wb, workload_rows)
    _write_blueprint(wb, blueprint)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _write_teacher_tt_summary(wb, teacher_data, teacher_info):
    ws = wb.create_sheet("Teacher TT Summary")
    ws.merge_cells("A1:K1")
    apply_cell(ws["A1"], "TEACHER TIMETABLE SUMMARY", HDR_FILL, HDR_FONT, CTR_ALIGN)
    for ci, h in enumerate(["Teacher", "Day"] + PERIODS, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = TTL_FILL
        cell.font = HDR_FONT
        cell.alignment = CTR_ALIGN
    row = 3
    for sheet_name, grid in teacher_data.items():
        info  = teacher_info[sheet_name]
        tname = f"{info['name']}-{info['id']}" if info['id'] else info['name']
        first = True
        for day in DAYS:
            tc = ws.cell(row=row, column=1, value=tname if first else None)
            tc.font = Font(bold=True, name="Arial", size=9)
            tc.alignment = CTR_ALIGN
            first = False
            ws.cell(row=row, column=2, value=day).fill = DAY_FILL
            ws.cell(row=row, column=2).font = DAY_FONT
            ws.cell(row=row, column=2).alignment = CTR_ALIGN
            for pi, period in enumerate(PERIODS, 3):
                val  = grid.get((day, period), "")
                cell = ws.cell(row=row, column=pi, value=val)
                cell.font = NRM_FONT
                cell.alignment = NRM_ALIGN
                cell.border = thin_border()
            row += 1
        ws.append([""])
        row += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 8
    for ci in range(3, 12):
        ws.column_dimensions[get_column_letter(ci)].width = 22


def _write_class_tt_summary(wb, class_tt):
    ws = wb.create_sheet("Class TT Summary")
    ws.merge_cells("A1:K1")
    apply_cell(ws["A1"], "CLASS TIMETABLE SUMMARY", HDR_FILL, HDR_FONT, CTR_ALIGN)
    for ci, h in enumerate(["Class", "Day"] + PERIODS, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = TTL_FILL
        cell.font = HDR_FONT
        cell.alignment = CTR_ALIGN
    row = 3
    for cls in sorted(class_tt.keys(), key=class_sort_key):
        grid  = class_tt[cls]
        first = True
        for day in DAYS:
            ws.cell(row=row, column=1, value=cls if first else None).font = Font(bold=True, name="Arial", size=9)
            ws.cell(row=row, column=1).alignment = CTR_ALIGN
            first = False
            ws.cell(row=row, column=2, value=day).fill = DAY_FILL
            ws.cell(row=row, column=2).font = DAY_FONT
            ws.cell(row=row, column=2).alignment = CTR_ALIGN
            for pi, period in enumerate(PERIODS, 3):
                val  = grid.get((day, period), "")
                cell = ws.cell(row=row, column=pi, value=val)
                cell.font = NRM_FONT
                cell.alignment = NRM_ALIGN
                cell.border = thin_border()
            row += 1
        ws.append([""])
        row += 1
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    for ci in range(3, 12):
        ws.column_dimensions[get_column_letter(ci)].width = 30


def _write_workload_summary(wb, workload_rows):
    ws = wb.create_sheet("Teacher Workload Summary")
    ws.merge_cells("A1:G1")
    apply_cell(ws["A1"], "TEACHER WORKLOAD SUMMARY", HDR_FILL, HDR_FONT, CTR_ALIGN)
    headers = ["Teacher Name", "Teacher ID", "Periods Allotted",
               "Unavail Periods", "Available Slots", "Workload %", "Status"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = TTL_FILL
        cell.font = HDR_FONT
        cell.alignment = CTR_ALIGN
    for ri, r in enumerate(workload_rows, 3):
        ws.cell(row=ri, column=1, value=r["Teacher Name"]).font = NRM_FONT
        ws.cell(row=ri, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=ri, column=2, value=r["Teacher ID"]).font = NRM_FONT
        ws.cell(row=ri, column=2).alignment = CTR_ALIGN
        ws.cell(row=ri, column=3, value=r["Periods Allotted"]).font = NRM_FONT
        ws.cell(row=ri, column=3).alignment = CTR_ALIGN
        # Unavail periods — highlight in amber if teacher has any
        uval = r.get("Unavail Periods", "None")
        uc = ws.cell(row=ri, column=4, value=uval)
        uc.font = NRM_FONT
        uc.alignment = CTR_ALIGN
        if uval != "None":
            uc.fill = PatternFill("solid", fgColor="FFF2CC")
            uc.font = Font(name="Arial", size=9, color="7F6000")
        ws.cell(row=ri, column=5, value=r.get("Available Slots", TOTAL_SLOTS)).font = NRM_FONT
        ws.cell(row=ri, column=5).alignment = CTR_ALIGN
        ws.cell(row=ri, column=6, value=f"{r['Workload %']:.2f}%").font = NRM_FONT
        ws.cell(row=ri, column=6).alignment = CTR_ALIGN
        status = r["Status"]
        sc = ws.cell(row=ri, column=7, value=status)
        sc.font = Font(bold=True, color=STATUS_COLORS.get(status, "000000"), name="Arial", size=9)
        sc.alignment = CTR_ALIGN
        for ci in range(1, 8):
            ws.cell(row=ri, column=ci).border = thin_border()
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14


def _write_unavailability_sheet(wb, unavail_rows):
    """Copy the Unavailability data as a clean formatted sheet."""
    ws = wb.create_sheet("Unavailability")

    ws.merge_cells("A1:C1")
    apply_cell(ws["A1"], "TEACHER UNAVAILABILITY", HDR_FILL, HDR_FONT, CTR_ALIGN)

    # Sub-header explaining the periods
    ws.merge_cells("A2:C2")
    ws["A2"].value = "Periods listed are unavailable for ALL days (P0=Assembly, P1-P8=Teaching periods)"
    ws["A2"].font  = Font(italic=True, name="Arial", size=9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Teacher Name", "Unavailable Periods", "Total Unavail Slots (× 6 days)"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.fill = TTL_FILL
        cell.font = HDR_FONT
        cell.alignment = CTR_ALIGN

    amber_fill = PatternFill("solid", fgColor="FFF2CC")
    amber_font = Font(name="Arial", size=9, color="7F6000")

    for ri, row in enumerate(unavail_rows, 4):
        tc = ws.cell(row=ri, column=1, value=row["teacher_key"])
        tc.font = NRM_FONT
        tc.alignment = Alignment(horizontal="left", vertical="center")

        pc = ws.cell(row=ri, column=2, value=", ".join(row["periods"]))
        pc.font = amber_font
        pc.fill = amber_fill
        pc.alignment = CTR_ALIGN

        total_unavail = len(row["periods"]) * len(DAYS)
        nc = ws.cell(row=ri, column=3, value=total_unavail)
        nc.font = Font(bold=True, name="Arial", size=9, color="CC0000")
        nc.alignment = CTR_ALIGN

        for ci in range(1, 4):
            ws.cell(row=ri, column=ci).border = thin_border()

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 28


def _write_blueprint(wb, blueprint):
    ws = wb.create_sheet("Blueprint")
    ws.merge_cells("A1:E1")
    apply_cell(ws["A1"], "TEACHING BLUEPRINT", HDR_FILL, HDR_FONT, CTR_ALIGN)
    for ci, h in enumerate(["Teacher Name", "Teacher ID", "Class", "Subject", "Periods"], 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = TTL_FILL
        cell.font = HDR_FONT
        cell.alignment = CTR_ALIGN
    row = 3
    for sheet_name, data in blueprint.items():
        info, counts = data["info"], data["counts"]
        if not counts:
            continue
        first = True
        for (cls, subj), cnt in sorted(counts.items()):
            ws.cell(row=row, column=1, value=info["name"] if first else "").font = NRM_FONT
            ws.cell(row=row, column=2, value=info["id"] if first else "").font = NRM_FONT
            first = False
            ws.cell(row=row, column=3, value=cls).font = NRM_FONT
            ws.cell(row=row, column=3).alignment = CTR_ALIGN
            ws.cell(row=row, column=4, value=subj).font = NRM_FONT
            ws.cell(row=row, column=4).alignment = CTR_ALIGN
            ws.cell(row=row, column=5, value=cnt).font = NRM_FONT
            ws.cell(row=row, column=5).alignment = CTR_ALIGN
            for ci in range(1, 6):
                ws.cell(row=row, column=ci).border = thin_border()
            row += 1
        total = sum(counts.values())
        ws.cell(row=row, column=4, value="TOTAL").font = Font(bold=True, name="Arial", size=9)
        ws.cell(row=row, column=5, value=total).font   = Font(bold=True, name="Arial", size=9)
        ws.cell(row=row, column=4).fill = SUM_FILL
        ws.cell(row=row, column=5).fill = SUM_FILL
        ws.cell(row=row, column=4).alignment = CTR_ALIGN
        ws.cell(row=row, column=5).alignment = CTR_ALIGN
        row += 1
        ws.append([""])
        row += 1
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 10


# ─────────────────────────────────────────────
#  STREAMLIT UI
# ─────────────────────────────────────────────
def main():
    st.title("📅 Timetable Sync & Conflict Checker")
    st.markdown("Upload your **Teacher Wise TT** to auto-generate Class TT + Summary Report and detect all conflicts.")
    st.divider()

    uploaded = st.file_uploader(
        "📂 Upload Teacher Wise TT (.xlsx)", type=["xlsx"],
        help="Upload the Teacher_Wise_TT file. All other files are auto-generated."
    )

    if not uploaded:
        st.info("👆 Upload your Teacher Wise TT file to get started.")
        with st.expander("ℹ️ What conflicts are detected?"):
            for code, desc in CONFLICT_DESCRIPTIONS.items():
                sev = "🔴" if code in ["C1","C2","C5","C8"] else "🟠" if code in ["C4","C6","C7","C10"] else "🔵"
                st.markdown(f"{sev} **{code}** — {desc}")
        return

    file_bytes = uploaded.read()

    with st.spinner("🔄 Parsing + analysing timetable..."):
        try:
            unavailability, unavail_rows, _ = parse_unavailability(file_bytes)
            teacher_data, teacher_info, teaching_summary = parse_teacher_tt(file_bytes)
        except Exception as e:
            st.error(f"❌ Error parsing file: {e}")
            return
        class_tt      = build_class_mapping(teacher_data, teacher_info)
        workload_rows = compute_workload(teacher_data, teacher_info, teaching_summary, unavailability)
        blueprint     = compute_blueprint(teacher_data, teacher_info)
        conflicts     = detect_conflicts(teacher_data, teacher_info)

    n_teachers = len(teacher_data)
    n_classes  = len(class_tt)
    n_critical = sum(1 for c in conflicts if c["severity"] == "CRITICAL")
    n_warning  = sum(1 for c in conflicts if c["severity"] == "WARNING")
    n_info     = sum(1 for c in conflicts if c["severity"] == "INFO")

    # Unavailability info banner
    if unavail_rows:
        with st.expander(f"📵 Unavailability loaded — {len(unavail_rows)} teachers have restricted periods", expanded=False):
            df_unavail = pd.DataFrame([
                {
                    "Teacher": r["teacher_key"],
                    "Unavailable Periods": ", ".join(r["periods"]),
                    "Unavail Slots (×6 days)": len(r["periods"]) * 6,
                    "Available Slots": TOTAL_SLOTS - len(r["periods"]) * 6,
                }
                for r in unavail_rows
            ])
            st.dataframe(df_unavail, use_container_width=True, hide_index=True)
            st.caption("Workload % is calculated against **Available Slots**, not total 54 slots.")
    if n_critical > 0:
        st.error(f"🔴 **{n_critical} critical conflict(s)** found — timetable needs fixing before use!")
    elif n_warning > 0:
        st.warning(f"🟠 **{n_warning} warning(s)** found — review before finalising.")
    else:
        st.success("✅ No critical conflicts! Timetable looks clean.")

    # Top metrics
    cols = st.columns(6)
    cols[0].metric("👨‍🏫 Teachers", n_teachers)
    cols[1].metric("🏫 Classes", n_classes)
    cols[2].metric("📋 Total Issues", len(conflicts))
    cols[3].metric("🔴 Critical", n_critical)
    cols[4].metric("🟠 Warnings", n_warning)
    cols[5].metric("🔵 Info", n_info)

    st.divider()

    # Tabs
    tab_conflict, tab_workload, tab_class, tab_blueprint, tab_teacher = st.tabs([
        f"⚠️ Conflicts & Errors ({len(conflicts)})",
        "📊 Workload",
        "🏫 Class TT Preview",
        "🗺️ Blueprint",
        "👨‍🏫 Teacher TT Preview",
    ])

    # ── CONFLICTS TAB ─────────────────────────────
    with tab_conflict:
        if not conflicts:
            st.success("🎉 No conflicts detected! Your timetable is clean.")
        else:
            col_f1, col_f2, col_f3 = st.columns([2, 2, 3])
            with col_f1:
                sev_filter = st.multiselect("Severity", ["CRITICAL", "WARNING", "INFO"],
                                            default=["CRITICAL", "WARNING", "INFO"])
            with col_f2:
                code_filter = st.multiselect("Conflict Type",
                                             sorted(set(c["conflict_id"] for c in conflicts)),
                                             default=sorted(set(c["conflict_id"] for c in conflicts)))
            with col_f3:
                search_text = st.text_input("🔍 Search teacher / class / keyword", "")

            filtered = [
                c for c in conflicts
                if c["severity"] in sev_filter
                and c["conflict_id"] in code_filter
                and (not search_text or any(
                    search_text.lower() in str(c.get(f, "") or "").lower()
                    for f in ["teacher", "cls", "title", "detail"]
                ))
            ]

            st.markdown(f"**Showing {len(filtered)} of {len(conflicts)} issues**")

            # Summary table
            summary_rows = []
            for c in filtered:
                summary_rows.append({
                    "Severity": c["severity"],
                    "Code": c["conflict_id"],
                    "Title": c["title"],
                    "Day": c.get("day") or "—",
                    "Period": c.get("period") or "—",
                    "Teacher": c.get("teacher") or "—",
                    "Class": c.get("cls") or "—",
                })
            if summary_rows:
                df_conf = pd.DataFrame(summary_rows)

                def highlight_sev(row):
                    colors = {"CRITICAL": "background-color: #FFE0E0",
                              "WARNING":  "background-color: #FFF3CD",
                              "INFO":     "background-color: #EAF4FF"}
                    return [colors.get(row["Severity"], "")] * len(row)

                st.dataframe(
                    df_conf.style.apply(highlight_sev, axis=1),
                    use_container_width=True, height=300
                )

            st.divider()
            st.subheader("Detailed Conflict Descriptions")

            sev_order = {"CRITICAL": 0, "WARNING": 1, "INFO": 2}
            for c in sorted(filtered, key=lambda x: sev_order.get(x["severity"], 3)):
                sev   = c["severity"]
                emoji = SEV_EMOJI[sev]
                with st.expander(f"{emoji} [{c['conflict_id']}] {c['title']}", expanded=(sev == "CRITICAL")):
                    ca, cb, cc = st.columns(3)
                    ca.markdown(f"**Severity:** <span style='color:{SEV_COLOR[sev]};font-weight:bold'>{sev}</span>",
                                unsafe_allow_html=True)
                    if c.get("day"):
                        cb.markdown(f"**When:** {c['day']} {c.get('period','')}")
                    if c.get("teacher"):
                        cc.markdown(f"**Teacher:** {c['teacher']}")
                    if c.get("cls"):
                        st.markdown(f"**Class:** {c['cls']}")
                    st.markdown(c["detail"])
                    st.caption(f"Rule: {CONFLICT_DESCRIPTIONS.get(c['conflict_id'], '')}")

    # ── WORKLOAD TAB ──────────────────────────────
    with tab_workload:
        st.subheader("Teacher Workload Summary")
        st.caption("⚠️ Workload % = Periods Allotted ÷ Available Slots (total 54 minus unavailable periods × 6 days)")

        df_wl = pd.DataFrame(workload_rows)
        df_wl["Workload %"] = df_wl["Workload %"].apply(lambda x: f"{x:.2f}%")

        status_filter = st.multiselect(
            "Filter by Status",
            ["Overloaded", "High Load", "Normal", "Low Load"],
            default=["Overloaded", "High Load", "Normal", "Low Load"]
        )
        df_show = df_wl[df_wl["Status"].isin(status_filter)]

        def color_status(val):
            c = {"Overloaded": "#FF4444", "High Load": "#FF9900",
                 "Normal": "#00AA44",     "Low Load": "#0088CC"}
            return f"color: {c.get(val,'black')}; font-weight: bold"

        def highlight_unavail(val):
            if val and val != "None":
                return "background-color: #FFF2CC; color: #7F6000"
            return ""

        st.dataframe(
            df_show.style
                .applymap(color_status, subset=["Status"])
                .applymap(highlight_unavail, subset=["Unavail Periods"]),
            use_container_width=True, height=420
        )
        st.bar_chart(df_wl["Status"].value_counts())

    # ── CLASS TT TAB ──────────────────────────────
    with tab_class:
        st.subheader("Class TT Preview")
        sel_class = st.selectbox("Select Class", sorted(class_tt.keys(), key=class_sort_key))
        if sel_class:
            grid = class_tt[sel_class]
            rows_preview = [{"Day": day, **{p: grid.get((day, p), "") for p in PERIODS}} for day in DAYS]
            st.dataframe(pd.DataFrame(rows_preview).set_index("Day"), use_container_width=True, height=280)
            cls_issues = [c for c in conflicts
                          if c.get("cls") == sel_class or sel_class in (c.get("detail") or "")]
            if cls_issues:
                st.warning(f"⚠️ {len(cls_issues)} issue(s) for class {sel_class}:")
                for c in cls_issues:
                    st.markdown(f"- {SEV_EMOJI[c['severity']]} **[{c['conflict_id']}]** {c['title']}")

    # ── BLUEPRINT TAB ─────────────────────────────
    with tab_blueprint:
        st.subheader("Teaching Blueprint")
        bp_rows = []
        for sheet_name, data in blueprint.items():
            info = data["info"]
            for (cls, subj), cnt in sorted(data["counts"].items()):
                bp_rows.append({"Teacher": info["name"], "ID": info["id"],
                                "Class": cls, "Subject": subj, "Periods": cnt})
        df_bp = pd.DataFrame(bp_rows)
        cb1, cb2 = st.columns(2)
        with cb1:
            sel_t = st.selectbox("Filter Teacher", ["All"] + sorted(df_bp["Teacher"].unique().tolist()))
        with cb2:
            sel_c = st.selectbox("Filter Class", ["All"] + sorted(set(df_bp["Class"].tolist()), key=class_sort_key))
        if sel_t != "All": df_bp = df_bp[df_bp["Teacher"] == sel_t]
        if sel_c != "All": df_bp = df_bp[df_bp["Class"] == sel_c]
        st.dataframe(df_bp, use_container_width=True, height=380)

    # ── TEACHER TT TAB ────────────────────────────
    with tab_teacher:
        st.subheader("Teacher TT Preview")
        sel_teacher = st.selectbox("Select Teacher", list(teacher_data.keys()))
        if sel_teacher:
            grid = teacher_data[sel_teacher]
            rows_preview = [{"Day": day, **{p: grid.get((day, p), "") for p in PERIODS}} for day in DAYS]
            st.dataframe(pd.DataFrame(rows_preview).set_index("Day"), use_container_width=True, height=280)
            info = teacher_info[sel_teacher]
            t_issues = [c for c in conflicts if c.get("teacher") and info["name"] in c["teacher"]]
            if t_issues:
                st.warning(f"⚠️ {len(t_issues)} issue(s) for {info['name']}:")
                for c in t_issues:
                    st.markdown(f"- {SEV_EMOJI[c['severity']]} **[{c['conflict_id']}]** {c['title']}")

    st.divider()

    # ── DOWNLOADS ─────────────────────────────────
    st.subheader("📥 Download Generated Files")
    with st.spinner("📝 Generating Excel files..."):
        class_tt_bytes = write_class_tt(class_tt, teacher_info)
        summary_bytes  = write_summary_report(teacher_data, teacher_info, class_tt, workload_rows, blueprint, unavail_rows)
        conflict_bytes = write_conflict_report(conflicts)

    dl1, dl2, dl3 = st.columns(3)
    with dl1:
        st.download_button("⬇️ Download Class Wise TT", data=class_tt_bytes,
                           file_name="Class_Wise_TT_Updated.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, type="primary")
        st.caption(f"{n_classes} class sheets")
    with dl2:
        st.download_button("⬇️ Download Summary Report", data=summary_bytes,
                           file_name="Summary_Report_Updated.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, type="primary")
        st.caption("Teacher TT + Class TT + Workload + Blueprint")
    with dl3:
        label = f"⬇️ Download Conflict Report ({len(conflicts)} issues)"
        st.download_button(label, data=conflict_bytes,
                           file_name="Conflict_Report.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True,
                           type="primary" if n_critical > 0 else "secondary")
        st.caption(f"🔴 {n_critical} Critical  🟠 {n_warning} Warnings  🔵 {n_info} Info")

    st.divider()
    st.caption("💡 Edit Teacher Wise TT → re-upload → all files + conflict report update instantly.")


if __name__ == "__main__":
    main()
